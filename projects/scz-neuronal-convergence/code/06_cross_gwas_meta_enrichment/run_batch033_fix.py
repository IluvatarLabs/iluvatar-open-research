#!/usr/bin/env python3
"""
Fix batch_033 failures:
1. Rebuild S-LDSC annotations on FULL plink SNP set (not baselineLD subset)
2. Fix D30.3 joint regression (gene TSS column names)
3. Fix D30.5 rank-rank overlap (MAGMA column names)
4. Fix D20 brain-expressed (ENSG→symbol mapping)
5. Fix D30.4 cross-disorder (column index)
"""

import pandas as pd
import numpy as np
import gzip
import json
import os
import subprocess
import time
from pathlib import Path
from scipy import stats
from scipy.stats import fisher_exact
from statsmodels.stats.multitest import multipletests
import statsmodels.api as sm
import warnings
warnings.filterwarnings('ignore')

PROJECT_ROOT = Path("/home/yuanz/Documents/GitHub/biomarvin_schizophrenia")
LDSC_BIN = "/home/yuanz/torchml/bin/ldsc.py"
OUTPUT_DIR = PROJECT_ROOT / "experiments/batch_033/output"
PLINK_DIR = PROJECT_ROOT / "data/ldsc/plink_format"
BASELINELD_DIR = PROJECT_ROOT / "data/ldsc/baselineLD"
WEIGHTS_DIR = PROJECT_ROOT / "data/ldsc/weights/1000G_Phase3_weights_hm3_no_MHC"
SUMSTATS = PROJECT_ROOT / "data/ldsc/PGC3_sumstats/PGC3_EUR_v2.sumstats.gz"
MARKERS_PATH = PROJECT_ROOT / "experiments/batch_009/data/markers.parquet"
GTEX_PATH = PROJECT_ROOT / "data/GTEx_v8_gene_median_tpm.gct.gz"
GENE_TSS_PATH = PROJECT_ROOT / "data/ldsc/gene_tss_grch37.csv"
BACKGROUND_SIZE = 20297

OUTPUT_DIR.mkdir(exist_ok=True, parents=True)

print("=" * 70)
print("batch_033 fix — Rebuilding S-LDSC annotations and fixing failures")
print("=" * 70)

# Load markers
markers_df = pd.read_parquet(MARKERS_PATH)
CELL_TYPES = {
    'Neurons': markers_df[markers_df['cell_type'] == 'Neurons']['gene'].unique(),
    'Oligodendrocytes': markers_df[markers_df['cell_type'] == 'Oligodendrocytes']['gene'].unique(),
    'Astrocytes': markers_df[markers_df['cell_type'] == 'Astrocytes']['gene'].unique(),
    'Oligodendrocyte progenitor cells': markers_df[markers_df['cell_type'] == 'Oligodendrocyte progenitor cells']['gene'].unique(),
}
markers = {ct: set(genes) for ct, genes in CELL_TYPES.items()}

# Load gene TSS for annotation building
gene_tss = pd.read_csv(GENE_TSS_PATH)
print(f"Gene TSS file: {len(gene_tss)} entries, columns: {gene_tss.columns.tolist()}")
# Filter to protein_coding
gene_tss_pc = gene_tss[gene_tss['biotype'] == 'protein_coding']
print(f"Protein-coding genes: {len(gene_tss_pc)}")

# Build TSS lookup: gene -> (chrom, tss_position)
# Need to handle non-standard chromosome names
WINDOW = 100_000  # 100kb

tss_lookup = {}
for _, row in gene_tss_pc.iterrows():
    gene = row['gene']
    chrom = str(row['chrom'])
    tss = int(row['tss'])
    # Normalize chromosome
    if not chrom.startswith('chr'):
        chrom_num = chrom
    else:
        chrom_num = chrom.replace('chr', '')
    tss_lookup[gene] = (chrom_num, tss)

print(f"TSS lookup for {len(tss_lookup)} genes")

# ============================================================================
# PART A: Rebuild S-LDSC annotations on FULL plink SNP set
# ============================================================================
print("\n## Part A: Rebuilding S-LDSC annotations on full plink SNP set")

annot_output = OUTPUT_DIR / "annotations_full"
annot_output.mkdir(exist_ok=True)

for chr in range(1, 23):
    bim_file = PLINK_DIR / f"chr{chr}_eur.bim"
    if not bim_file.exists():
        print(f"  chr{chr}: bim file not found, skipping")
        continue

    # Read plink BIM file
    bim = pd.read_csv(bim_file, sep='\t', header=None,
                       names=['chrom', 'snp', 'cm', 'pos', 'a1', 'a2'])

    # Build annotation columns
    neuronal = np.zeros(len(bim), dtype=int)
    oligodendrocyte = np.zeros(len(bim), dtype=int)
    astrocyte = np.zeros(len(bim), dtype=int)
    opc = np.zeros(len(bim), dtype=int)

    # For each marker gene, find SNPs in TSS ±100kb window
    for ct_name, ct_markers in CELL_TYPES.items():
        target_col = {
            'Neurons': neuronal,
            'Oligodendrocytes': oligodendrocyte,
            'Astrocytes': astrocyte,
            'Oligodendrocyte progenitor cells': opc,
        }[ct_name]

        for gene in ct_markers:
            if gene not in tss_lookup:
                continue
            gchr, gtss = tss_lookup[gene]
            # Handle chr mismatch
            try:
                gchr_int = int(gchr)
            except ValueError:
                continue
            if gchr_int != chr:
                continue

            # Mark SNPs in window
            in_window = (bim['pos'] >= gtss - WINDOW) & (bim['pos'] <= gtss + WINDOW)
            target_col[in_window.values] = 1

    # Create annotation dataframe
    annot_df = pd.DataFrame({
        'CHR': bim['chrom'],
        'BP': bim['pos'],
        'SNP': bim['snp'],
        'CM': bim['cm'],
        'base': 1,  # Required for --overlap-annot
        'neuronal': neuronal,
        'oligodendrocyte': oligodendrocyte,
        'astrocyte': astrocyte,
        'OPC': opc,
    })

    # Write gzipped annotation file
    out_file = annot_output / f"celltype.{chr}.annot.gz"
    with gzip.open(out_file, 'wt') as f:
        annot_df.to_csv(f, sep='\t', index=False)

    n_neur = neuronal.sum()
    n_olig = oligodendrocyte.sum()
    n_ast = astrocyte.sum()
    n_opc = opc.sum()
    print(f"  chr{chr}: {len(bim)} SNPs, neuronal={n_neur}, oligo={n_olig}, "
          f"astro={n_ast}, OPC={n_opc}")

print(f"\nAnnotations built in {annot_output}")

# ============================================================================
# PART B: Compute LD scores from new annotations (parallel)
# ============================================================================
print("\n## Part B: Computing cell-type LD scores (22 chromosomes in parallel)")

ld_dir = OUTPUT_DIR / "celltype_ld_scores_v2"
ld_dir.mkdir(exist_ok=True)

hm3_snps = WEIGHTS_DIR / "hm3_snps.txt"
processes = []
for chr in range(1, 23):
    annot_file = annot_output / f"celltype.{chr}.annot.gz"
    if not annot_file.exists():
        continue

    cmd = (
        f"python3 {LDSC_BIN} "
        f"--l2 "
        f"--bfile {PLINK_DIR}/chr{chr}_eur "
        f"--ld-wind-cm 1.0 "
        f"--annot {annot_file} "
        f"--out {ld_dir}/celltype.{chr} "
        f"--print-snps {hm3_snps}"
    )
    p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    processes.append((chr, p))

print(f"  Launched {len(processes)} parallel LD score computations...")
for chr, p in processes:
    stdout, stderr = p.communicate()
    if p.returncode != 0:
        err = stderr.decode()[-200:]
        print(f"  chr{chr} FAILED: {err}")
    else:
        # Check output
        ldscore_file = ld_dir / f"celltype.{chr}.l2.ldscore.gz"
        if ldscore_file.exists():
            n_lines = sum(1 for _ in gzip.open(ldscore_file)) - 1
            print(f"  chr{chr}: OK ({n_lines} SNPs with LD scores)")
        else:
            print(f"  chr{chr}: completed but no ldscore file")

# ============================================================================
# PART C: Run S-LDSC partitioned heritability
# ============================================================================
print("\n## Part C: Running S-LDSC partitioned heritability")

celltype_ld = str(ld_dir) + "/celltype."
baseline_ld = str(BASELINELD_DIR) + "/baselineLD."
weights_ld = str(WEIGHTS_DIR) + "/weights.hm3_noMHC."

cmd = (
    f"python3 {LDSC_BIN} "
    f"--h2 {SUMSTATS} "
    f"--ref-ld-chr {celltype_ld},{baseline_ld} "
    f"--w-ld-chr {weights_ld} "
    f"--overlap-annot "
    f"--out {OUTPUT_DIR}/celltype_partitioned_v2"
)

print(f"  Running partitioned heritability...")
result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=3600)

if result.returncode != 0:
    print(f"  FAILED: {result.stderr[-500:]}")
    # Try without --overlap-annot
    cmd2 = (
        f"python3 {LDSC_BIN} "
        f"--h2 {SUMSTATS} "
        f"--ref-ld-chr {celltype_ld},{baseline_ld} "
        f"--w-ld-chr {weights_ld} "
        f"--out {OUTPUT_DIR}/celltype_partitioned_v2"
    )
    print("  Retrying without --overlap-annot...")
    result = subprocess.run(cmd2, shell=True, capture_output=True, text=True, timeout=3600)

results_file = OUTPUT_DIR / "celltype_partitioned_v2.results"
if results_file.exists():
    d32_results = pd.read_csv(results_file, sep='\t')
    print("\n  S-LDSC Cell-Type Results:")
    # Filter to cell-type annotations (L2_0 = from first prefix = celltype)
    ct_results = d32_results[d32_results['Category'].str.contains('L2_0')]
    print(ct_results[['Category', 'Coefficient', 'Coefficient_std_error',
                       'Coefficient_p_value', 'Enrichment']].to_string(index=False))

    with open(OUTPUT_DIR / "d32_sldsc_results.json", 'w') as f:
        json.dump({
            'status': 'success',
            'results_file': str(results_file),
            'cell_type_results': ct_results.to_dict('records'),
            'all_results': d32_results.to_dict('records')[:20]  # First 20 for size
        }, f, indent=2, default=str)
else:
    print(f"  Results file not found")
    print(f"  stdout: {result.stdout[-500:]}")
    print(f"  stderr: {result.stderr[-500:]}")

# ============================================================================
# PART D: Fix D30.3 — Joint Logistic Regression
# ============================================================================
print("\n" + "=" * 70)
print("PART D: D30.3 — Conditional Analysis (fixed)")
print("=" * 70)

# Load gene lists
pardinas_df = pd.read_parquet(PROJECT_ROOT / "experiments/batch_008/data/gwas_genes.parquet")
pardinas_genes = set(pardinas_df['hgnc_symbol'].unique())

pgc3_df = pd.read_csv(PROJECT_ROOT / "experiments/batch_025/data/pgc3_gene_list.csv")
pgc3_genes = set(pgc3_df['gene'].unique())

# Build feature matrix using gene TSS file (correct column: 'gene')
gene_universe = set(gene_tss_pc['gene'].unique())
print(f"  Gene universe: {len(gene_universe)} protein-coding genes")

# Load pathway gene sets
try:
    import gseapy as gp
    from gseapy import Msigdb
    kegg = Msigdb().get_gmt(category='c2.cp.kegg', dbver='2023.1.Hs')
    tlr_genes = set(kegg.get('KEGG_TOLL_LIKE_RECEPTOR_SIGNALING_PATHWAY', []))
except:
    tlr_genes = set()

try:
    import decoupler as dc
    dor = dc.op.dorothea(organism='human')
    nfkb_genes = set(dor[dor['source'].isin(['RELA', 'NFKB1'])]['target'].unique())
except:
    nfkb_genes = set()

# Load SynGO from PGC3 table
import openpyxl
wb = openpyxl.load_workbook(PROJECT_ROOT / "data/19426775/scz2022-Extended-Data-Table1.xlsx", read_only=True)
ws = wb['ST12 all criteria']
rows = list(ws.iter_rows(values_only=True))
synGo_genes = set()
for row in rows[1:]:
    if row[3] == 'protein_coding' and row[23] == 'YES':
        synGo_genes.add(row[2])

# Build features
features = []
for gene in gene_universe:
    features.append({
        'gene': gene,
        'is_SCZ': 1 if gene in pardinas_genes else 0,
        'neuronal': 1 if gene in markers.get('Neurons', set()) else 0,
        'oligodendrocyte': 1 if gene in markers.get('Oligodendrocytes', set()) else 0,
        'astrocyte': 1 if gene in markers.get('Astrocytes', set()) else 0,
        'OPC': 1 if gene in markers.get('Oligodendrocyte progenitor cells', set()) else 0,
        'SynGO': 1 if gene in synGo_genes else 0,
        'TLR': 1 if gene in tlr_genes else 0,
        'NFKB': 1 if gene in nfkb_genes else 0,
    })

df = pd.DataFrame(features)
print(f"  Feature matrix: {len(df)} genes, {df['is_SCZ'].sum()} SCZ genes")

predictors = ['neuronal', 'oligodendrocyte', 'astrocyte', 'OPC', 'SynGO', 'TLR', 'NFKB']
X = df[predictors]
y = df['is_SCZ']

try:
    X_const = sm.add_constant(X)
    model = sm.Logit(y, X_const).fit(method='lbfgs', maxiter=1000, disp=0)
    print("\n  Logistic Regression Results:")
    print(model.summary2().tables[1].to_string())

    # VIF
    from statsmodels.stats.outliers_influence import variance_inflation_factor
    vif_data = []
    for i, p in enumerate(predictors):
        vif = variance_inflation_factor(X.values, i)
        vif_data.append({'predictor': p, 'VIF': vif})
        print(f"  VIF({p}) = {vif:.2f}")

    coef_df = model.summary2().tables[1]
    regression_results = {
        'n_genes': len(df),
        'n_SCZ': int(df['is_SCZ'].sum()),
        'predictors': predictors,
        'coefficients': {
            p: {
                'coef': float(coef_df.loc[p, 'Coef.']),
                'p': float(coef_df.loc[p, 'P>|z|']),
                'OR': float(np.exp(coef_df.loc[p, 'Coef.'])),
            }
            for p in predictors if p in coef_df.index
        },
        'VIF': {v['predictor']: float(v['VIF']) for v in vif_data},
        'pseudo_R2': float(model.prsquared),
        'note': 'Conditional analysis — NOT independent evidence beyond Fisher exact'
    }
except Exception as e:
    print(f"  Regression failed: {e}")
    regression_results = {'status': 'failed', 'error': str(e)}

with open(OUTPUT_DIR / "d30_3_joint_regression.json", 'w') as f:
    json.dump(regression_results, f, indent=2, default=str)

# ============================================================================
# PART E: Fix D30.5 — Rank-Rank Overlap
# ============================================================================
print("\n" + "=" * 70)
print("PART E: D30.5 — Rank-Rank Overlap (fixed)")
print("=" * 70)

# Try different MAGMA file locations
magma_file = None
for path in [
    PROJECT_ROOT / "experiments/batch_026/gene_level_pgc3.tsv",
    PROJECT_ROOT / "experiments/batch_028/data/gene_level_pgc3.tsv",
]:
    if path.exists():
        magma_file = path
        break

if magma_file:
    magma_df = pd.read_csv(magma_file, sep='\t')
    print(f"  MAGMA results: {len(magma_df)} genes")
    print(f"  Columns: {magma_df.columns.tolist()}")

    # Find the p-value column
    p_col = None
    for col in ['P', 'p_value', 'stouffer_p', 'min_p', 'pval']:
        if col in magma_df.columns:
            p_col = col
            break

    if p_col:
        # Create MAGMA rank (lower p = higher rank)
        magma_df['magma_rank'] = magma_df[p_col].rank(ascending=True)
        neuronal_markers = markers['Neurons']
        magma_df['is_neuronal'] = magma_df['gene'].isin(neuronal_markers).astype(int)

        rho, p = stats.spearmanr(magma_df['magma_rank'], magma_df['is_neuronal'])
        print(f"\n  Spearman ρ(MAGMA rank, neuronal) = {rho:.4f}, p = {p:.4e}")

        # Top-N enrichment
        top_n = [50, 100, 200, 500]
        enrichment_at_top = []
        for n in top_n:
            top_genes = set(magma_df.nsmallest(n, p_col)['gene'])
            r = fisher_enrichment(neuronal_markers, top_genes, len(magma_df),
                                  label=f"Top-{n} MAGMA × Neuronal")
            enrichment_at_top.append({
                'n': n, 'OR': r['odds_ratio'], 'p': r['p_value'], 'k': r['overlap']
            })
            print(f"  Top-{n} MAGMA genes × Neuronal: OR={r['odds_ratio']:.2f}, "
                  f"p={r['p_value']:.4f}, k={r['overlap']}")
    else:
        print(f"  No p-value column found in MAGMA file")
        rho, p = None, None
        enrichment_at_top = []
else:
    print("  MAGMA gene-level results not found")
    rho, p = None, None
    enrichment_at_top = []

rro_results = {
    'spearman_rho': float(rho) if rho is not None else None,
    'spearman_p': float(p) if p is not None else None,
    'enrichment_at_top': enrichment_at_top,
}

with open(OUTPUT_DIR / "d30_5_rank_rank.json", 'w') as f:
    json.dump(rro_results, f, indent=2, default=str)

# ============================================================================
# PART F: Fix D20 — Brain-Expressed Background
# ============================================================================
print("\n" + "=" * 70)
print("PART F: D20 — Brain-Expressed Background (fixed)")
print("=" * 70)

gtex_df = pd.read_csv(GTEX_PATH, sep='\t', skiprows=2)
brain_cols = [c for c in gtex_df.columns if 'Brain' in c]
print(f"  Brain tissues ({len(brain_cols)}): {brain_cols}")

gtex_df['brain_max_tpm'] = gtex_df[brain_cols].max(axis=1)
brain_expressed_all = gtex_df[gtex_df['brain_max_tpm'] > 1]['Name'].unique()
print(f"  Brain-expressed genes (all biotypes, TPM>1): {len(brain_expressed_all)}")

# Map ENSG IDs to gene symbols using GTEx Description column
# GTEx has both Name (ENSG ID) and Description (gene symbol)
gtex_symbol_map = dict(zip(gtex_df['Name'], gtex_df['Description']))
brain_expressed_symbols = set()
for gid in brain_expressed_all:
    gid_clean = gid.split('.')[0]
    symbol = gtex_symbol_map.get(gid, None)
    if symbol:
        brain_expressed_symbols.add(symbol)
    else:
        brain_expressed_symbols.add(gid_clean)

# Filter to protein-coding only using gene TSS file
gene_tss_symbols = set(gene_tss_pc['gene'].unique())
brain_expressed_pc = brain_expressed_symbols & gene_tss_symbols
print(f"  Brain-expressed protein-coding genes: {len(brain_expressed_pc)}")

# Run enrichment with brain-expressed background
neuronal_markers = markers['Neurons']
neuronal_markers_brain = neuronal_markers & brain_expressed_pc

pardinas_genes_set = pardinas_genes
pgc3_genes_set = pgc3_genes

d20_results = {}
for gwas_name, gwas_genes in [('Pardiñas', pardinas_genes_set), ('PGC3', pgc3_genes_set)]:
    # Standard Entrez background
    r_std = fisher_enrichment(neuronal_markers, gwas_genes, BACKGROUND_SIZE,
                              label=f"{gwas_name} × Neuronal (Entrez)")

    # Brain-expressed background
    gwas_brain = gwas_genes & brain_expressed_pc
    N_brain = len(brain_expressed_pc)

    r_brain = fisher_enrichment(neuronal_markers_brain, gwas_brain, N_brain,
                                 label=f"{gwas_name} × Neuronal (brain bg)")

    print(f"\n  {gwas_name}:")
    print(f"    Entrez bg (N={BACKGROUND_SIZE}):    OR={r_std['odds_ratio']:.2f}, "
          f"p={r_std['p_value']:.4f}, k={r_std['overlap']}")
    print(f"    Brain-expressed bg (N={N_brain}): OR={r_brain['odds_ratio']:.2f}, "
          f"p={r_brain['p_value']:.4f}, k={r_brain['overlap']}")

    d20_results[gwas_name] = {
        'standard': r_std,
        'brain_expressed': r_brain,
        'brain_bg_size': N_brain,
        'brain_markers_in_bg': len(neuronal_markers_brain),
        'brain_gwas_in_bg': len(gwas_brain)
    }

# Also test oligodendrocytes
oligo_markers = markers['Oligodendrocytes']
oligo_brain = oligo_markers & brain_expressed_pc
for gwas_name, gwas_genes in [('Pardiñas', pardinas_genes_set)]:
    gwas_brain = gwas_genes & brain_expressed_pc
    r_oligo = fisher_enrichment(oligo_brain, gwas_brain, len(brain_expressed_pc),
                                 label=f"{gwas_name} × Oligo (brain bg)")
    print(f"\n  {gwas_name} × Oligodendrocyte (brain bg): OR={r_oligo['odds_ratio']:.2f}, "
          f"p={r_oligo['p_value']:.4f}")
    d20_results[f"{gwas_name}_oligo"] = r_oligo

with open(OUTPUT_DIR / "d20_brain_background.json", 'w') as f:
    json.dump(d20_results, f, indent=2, default=str)

# ============================================================================
# PART G: Fix D30.4 — Cross-Disorder (check column indices)
# ============================================================================
print("\n" + "=" * 70)
print("PART G: D30.4 — Cross-Disorder Conditional (fixed)")
print("=" * 70)

wb = openpyxl.load_workbook(PROJECT_ROOT / "data/19426775/scz2022-Extended-Data-Table1.xlsx", read_only=True)
ws = wb['ST12 all criteria']
rows = list(ws.iter_rows(values_only=True))
header = rows[0]

# Print column names with indices to debug
for i, h in enumerate(header):
    if h and ('SCHEMA' in str(h) or 'ASD' in str(h) or 'DDD' in str(h)):
        print(f"  Col {i}: {h} — first value: {rows[1][i] if rows[1][i] else 'None'}")

# Count cross-disorder genes with different YES indicators
schema_genes = set()
asd_genes = set()
ddd_genes = set()
for row in rows[1:]:
    if row[3] != 'protein_coding' or not row[2]:
        continue
    gene = row[2]
    # Check all columns that might contain SCHEMA/ASD/DDD
    for i, h in enumerate(header):
        if h and 'SCHEMA' in str(h) and row[i] == 'YES':
            schema_genes.add(gene)
        if h and 'ASD' in str(h) and row[i] == 'YES':
            asd_genes.add(gene)
        if h and 'DDD' in str(h) and row[i] == 'YES':
            ddd_genes.add(gene)

cross_disorder = schema_genes | asd_genes | ddd_genes
print(f"\n  SCHEMA: {len(schema_genes)} genes: {sorted(schema_genes)}")
print(f"  ASD: {len(asd_genes)} genes: {sorted(asd_genes)}")
print(f"  DDD: {len(ddd_genes)} genes: {sorted(ddd_genes)}")
print(f"  Union: {len(cross_disorder)} cross-disorder genes")

# Run conditional test on Pardiñas (adequate power)
if cross_disorder:
    for gwas_name, gwas_genes in [('Pardiñas', pardinas_genes_set), ('PGC3', pgc3_genes_set)]:
        r_orig = fisher_enrichment(neuronal_markers, gwas_genes, BACKGROUND_SIZE,
                                    label=f"{gwas_name} × Neuronal (original)")
        gwas_cond = gwas_genes - cross_disorder
        r_cond = fisher_enrichment(neuronal_markers, gwas_cond, BACKGROUND_SIZE,
                                    label=f"{gwas_name} × Neuronal (minus cross-disorder)")
        print(f"\n  {gwas_name}:")
        print(f"    Original:         OR={r_orig['odds_ratio']:.2f}, p={r_orig['p_value']:.4f}, k={r_orig['overlap']}/{len(gwas_genes)}")
        print(f"    -cross-disorder:  OR={r_cond['odds_ratio']:.2f}, p={r_cond['p_value']:.4f}, k={r_cond['overlap']}/{len(gwas_cond)}")
else:
    print("  No cross-disorder genes found — checking column values...")
    # Debug: show all values for SCHEMA/ASD/DDD columns
    for i, h in enumerate(header):
        if h and any(x in str(h) for x in ['SCHEMA', 'ASD', 'DDD']):
            vals = set(row[i] for row in rows[1:])
            print(f"  Col {i} ({h}): unique values = {vals}")

print("\n" + "=" * 70)
print("batch_033 fix COMPLETE")
print("=" * 70)
